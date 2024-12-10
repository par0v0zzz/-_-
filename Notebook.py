import sys
import sqlite3
import openpyxl
from openpyxl.styles import Font  # Импортируем Font из openpyxl
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton, QLabel, QListWidget, QInputDialog, QMessageBox, QLineEdit
)


class Database:
    def __init__(self, db_name):
        self.connection = sqlite3.connect(db_name)
        self.create_tables()

    def create_tables(self):
        with self.connection:
            self.connection.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE,
                    password TEXT
                )
            """)
            self.connection.execute("""
                CREATE TABLE IF NOT EXISTS notes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER,
                    content TEXT,
                    category TEXT,
                    tags TEXT,
                    FOREIGN KEY (user_id) REFERENCES users (id)
                )
            """)
            self.connection.execute("""
                CREATE TABLE IF NOT EXISTS tasks (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER,
                    content TEXT,
                    due_date TEXT,
                    priority INTEGER,
                    completed INTEGER DEFAULT 0,
                    FOREIGN KEY (user_id) REFERENCES users (id)
                )
            """)

    def add_user(self, username, password):
        with self.connection:
            self.connection.execute("INSERT INTO users (username, password) VALUES (?, ?)", (username, password))

    def validate_user(self, username, password):
        cursor = self.connection.cursor()
        cursor.execute("SELECT id FROM users WHERE username = ? AND password = ?", (username, password))
        return cursor.fetchone()

    def add_note(self, user_id, content, category, tags):
        with self.connection:
            self.connection.execute("INSERT INTO notes (user_id, content, category, tags) VALUES (?, ?, ?, ?)",
                                    (user_id, content, category, tags))

    def get_notes(self, user_id):
        cursor = self.connection.cursor()
        cursor.execute("SELECT content, category, tags FROM notes WHERE user_id = ?", (user_id,))
        return cursor.fetchall()

    def delete_note(self, user_id, content):
        with self.connection:
            self.connection.execute("DELETE FROM notes WHERE user_id = ? AND content = ?", (user_id, content))

    def add_task(self, user_id, content, due_date, priority):
        with self.connection:
            self.connection.execute("INSERT INTO tasks (user_id, content, due_date, priority) VALUES (?, ?, ?, ?)",
                                    (user_id, content, due_date, priority))

    def get_tasks(self, user_id):
        cursor = self.connection.cursor()
        cursor.execute("SELECT content, due_date, priority, completed FROM tasks WHERE user_id = ?", (user_id,))
        return cursor.fetchall()

    def mark_task_completed(self, user_id, content):
        with self.connection:
            self.connection.execute("UPDATE tasks SET completed = 1 WHERE user_id = ? AND content = ?",
                                    (user_id, content))


class LoginWindow(QWidget):
    def __init__(self, notebook_app):
        super().__init__()
        self.notebook_app = notebook_app
        self.db = Database("users.db")  # Подключение к базе данных
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Вход")
        self.setGeometry(100, 100, 300, 200)

        layout = QVBoxLayout()

        self.label = QLabel("Введите логин и пароль", self)
        layout.addWidget(self.label)

        self.username_input = QLineEdit(self)
        self.username_input.setPlaceholderText("Логин")
        layout.addWidget(self.username_input)

        self.password_input = QLineEdit(self)
        self.password_input.setPlaceholderText("Пароль")
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addWidget(self.password_input)

        login_button = QPushButton("Войти", self)
        login_button.clicked.connect(self.login)
        layout.addWidget(login_button)

        register_button = QPushButton("Регистрация", self)
        register_button.clicked.connect(self.open_registration)
        layout.addWidget(register_button)

        self.setLayout(layout)

    def login(self):
        """Авторизация пользователя"""
        username = self.username_input.text().strip()
        password = self.password_input.text().strip()

        if not username or not password:
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, заполните логин и пароль.")
            return

        user = self.db.validate_user(username, password)
        if user:
            self.notebook_app.user_id = user[0]  # Сохраните ID пользователя
            self.notebook_app.load_notes()  # Загрузите заметки пользователя
            self.notebook_app.show()
            self.close()
        else:
            QMessageBox.warning(self, "Ошибка", "Неверный логин или пароль.")

    def open_registration(self):
        self.registration_window = RegistrationWindow(self)
        self.registration_window.show()
        self.close()


class RegistrationWindow(QWidget):
    def __init__(self, login_window):
        super().__init__()
        self.login_window = login_window
        self.db = Database("users.db")  # Подключение к базе данных
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Регистрация")
        self.setGeometry(100, 100, 300, 200)

        layout = QVBoxLayout()

        self.label = QLabel("Введите логин и пароль для регистрации", self)
        layout.addWidget(self.label)

        self.username_input = QLineEdit(self)
        self.username_input.setPlaceholderText("Логин")
        layout.addWidget(self.username_input)

        self.password_input = QLineEdit(self)
        self.password_input.setPlaceholderText("Пароль")
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addWidget(self.password_input)

        register_button = QPushButton("Зарегистрироваться", self)
        register_button.clicked.connect(self.register)
        layout.addWidget(register_button)

        back_button = QPushButton("Назад", self)
        back_button.clicked.connect(self.go_back)
        layout.addWidget(back_button)

        self.setLayout(layout)

    def register(self):
        username = self.username_input.text().strip()
        password = self.password_input.text().strip()

        if not username or not password:
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, заполните логин и пароль.")
            return

        try:
            self.db.add_user(username, password)
            QMessageBox.information(self, "Успех", "Вы успешно зарегистрированы!")
            self.login_window.show()
            self.close()
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Ошибка", "Пользователь с таким именем уже существует.")

    def go_back(self):
        self.login_window.show()
        self.close()


class NotebookApp(QWidget):
    def __init__(self):
        super().__init__()
        self.user_id = None  # ID текущего пользователя
        self.db = Database("users.db")  # Подключение к базе данных

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Записная книга")
        self.setGeometry(100, 100, 400, 500)
        layout = QVBoxLayout()

        self.label = QLabel("Записная книга", self)
        layout.addWidget(self.label)

        self.entries_list = QListWidget(self)
        layout.addWidget(self.entries_list)

        add_note_button = QPushButton("Добавить заметку", self)
        add_note_button.clicked.connect(self.add_note)
        layout.addWidget(add_note_button)

        add_task_button = QPushButton("Добавить задачу", self)
        add_task_button.clicked.connect(self.add_task)
        layout.addWidget(add_task_button)

        view_tasks_button = QPushButton("Просмотреть задачи", self)
        view_tasks_button.clicked.connect(self.view_tasks)
        layout.addWidget(view_tasks_button)

        delete_note_button = QPushButton("Удалить заметку", self)
        delete_note_button.clicked.connect(self.delete_note)
        layout.addWidget(delete_note_button)

        create_report_button = QPushButton("Создать отчет", self)
        create_report_button.clicked.connect(self.export_to_excel)
        layout.addWidget(create_report_button)

        exit_button = QPushButton("Выход", self)
        exit_button.clicked.connect(self.close)
        layout.addWidget(exit_button)

        self.setLayout(layout)

    def add_note(self):
        content, ok = QInputDialog.getText(self, "Добавить заметку", "Введите текст заметки:")
        if ok and content:
            category, ok_category = QInputDialog.getText(self, "Категория", "Введите категорию:")
            if ok_category:
                tags, ok_tags = QInputDialog.getText(self, "Теги", "Введите теги (через запятую):")
                if ok_tags:
                    self.db.add_note(self.user_id, content, category, tags)
                    self.load_notes()

    def load_notes(self):
         if self.user_id is not None:
            notes = self.db.get_notes(self.user_id)
            self.entries_list.clear()
            for note in notes:
                self.entries_list.addItem(f"Заметка: {note[0]}, Категория: {note[1]}, Теги: {note[2]}")

    def delete_note(self):
        if self.entries_list.selectedItems():
            selected_item = self.entries_list.currentItem().text()
            content = selected_item.split(",")[0].replace("Заметка: ", "")
            self.db.delete_note(self.user_id, content)
            self.load_notes()
            QMessageBox.information(self, "Успех", "Заметка удалена!")
        else:
            QMessageBox.warning(self, "Ошибка", "Выберите заметку для удаления.")

    def add_task(self):
        content, ok = QInputDialog.getText(self, "Добавить задачу", "Введите текст задачи :")
        if ok and content:
            due_date, ok_date = QInputDialog.getText(self, "Срок выполнения",
                                                     "Введите срок выполнения (YYYY-MM-DD HH:MM):")
            if ok_date:
                priority, ok_priority = QInputDialog.getInt(self, "Приоритет", "Введите приоритет (1-5):", 1, 1, 5)
                if ok_priority:
                    self.db.add_task(self.user_id, content, due_date, priority)
                    QMessageBox.information(self, "Успех", "Задача добавлена!")

    def view_tasks(self):
        tasks = self.db.get_tasks(self.user_id)
        if tasks:
            task_list = "\n".join(
                f"Задача: {task[0]}, Срок: {task[1]}, Приоритет: {task[2]}, Выполнена: {'Да' if task[3] else 'Нет'}" for
                task in tasks)
            QMessageBox.information(self, "Задачи", task_list)
        else:
            QMessageBox.information(self, "Задачи", "Нет задач.")

    def export_to_excel(self):
        try:
            conn = sqlite3.connect("users.db")
            cursor = conn.cursor()

            cursor.execute("SELECT user_id, content, category, tags FROM notes")
            user_notes = cursor.fetchall()
            if not user_notes:
                QMessageBox.warning(self, "Пользователь не заводил заметки", "Заметки отсутствуют, ошибка выгрузки.")
                return

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Отчет"

            headers = ["ФИО пользователя", "Заметки", "Категория", "Теги"]
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(name="Arial", size=12, bold=True)  # Используйте правильный импорт для Font

            for row_num, (user_id, content, category, tags) in enumerate(user_notes, start=2):
                cursor.execute("SELECT username FROM users WHERE id = ?", (user_id,))
                username_row = cursor.fetchone()
                username = username_row[0] if username_row else "Unknown User"
                sheet.cell(row=row_num, column=1, value=username)  # ФИО
                sheet.cell(row=row_num, column=2, value=content)  # Заметки
                sheet.cell(row=row_num, column=3, value=category)  # Категория
                sheet.cell(row=row_num, column=4, value=tags)  # Теги

            # Сохранение файла
            output_path = "Отчет_Пользователи.xlsx"
            workbook.save(output_path)
            QMessageBox.information(self, "Успех", f"Данные успешно выгружены в файл {output_path}")

            # Открываем файл для проверки
            os.startfile(output_path)

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка выгрузки данных: {e}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {e}")

        finally:
            conn.close()


# Пример использования альтернативного подхода для управления шрифтами
class CustomFont:
    def __init__(self, name, size, bold=False, italic=False):
        self.name = name
        self.size = size
        self.bold = bold
        self.italic = italic


# Создание экземпляра CustomFont
my_font = CustomFont(name="Arial", size=12, bold=True)


# Использование my_font в вашем коде
def main():
    app = QApplication(sys.argv)

    # Создаем экземпляр основного окна записной книги
    notebook_app = NotebookApp()

    # Создаем окно для входа
    login_window = LoginWindow(notebook_app)
    login_window.show()

    # Запускаем приложение
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
